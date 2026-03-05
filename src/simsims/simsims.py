import time
import os
import random
import matplotlib.pyplot as plt
import platform
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

next_worker_id = 1
next_food_id = 1
next_product_id = 1


def get_next_worker_id():
    global next_worker_id
    worker_id = next_worker_id
    next_worker_id += 1
    return worker_id


def get_next_food_id():
    global next_food_id
    food_id = next_food_id
    next_food_id += 1
    return food_id


def get_next_product_id():
    global next_product_id
    product_id = next_product_id
    next_product_id += 1
    return product_id


class Worker:
    def __init__(self, worker_id, life_force=100):
        self.worker_id = worker_id
        self.life_force = life_force
        self.is_alive = True

    def adjust_life_force(self, amount):
        self.life_force += amount
        if self.life_force > 100:
            self.life_force = 100
        if self.life_force <= 0:
            self.is_alive = False

    def __repr__(self):
        if self.is_alive:
            return f"Worker-: {self.worker_id}(HP:{self.life_force})"
        else:
            return f"Worker-: {self.worker_id}(DEAD)"


class Food:
    def __init__(self, food_id, nutrition_value=30):
        self.food_id = food_id
        self.nutrition_value = nutrition_value

    def __repr__(self):
        return f"Food-{self.food_id}(+{self.nutrition_value}HP)"


class Product:
    def __init__(self, product_id):
        self.product_id = product_id

    def __repr__(self):
        return f"Product-{self.product_id}"


class Place:
    def __init__(self, name):
        self.name = name
        self.items = []

    def is_empty(self):
        return len(self.items) == 0

    def add_item(self, item):
        self.items.append(item)

    def remove_item(self):
        if not self.is_empty():
            return self.items.pop(0)
        return None

    def __repr__(self):
        return f"{self.name}: {self.items}"


class Barracks(Place):
    def __init__(self, name):
        super().__init__(name)

    def add_item(self, item):
        if isinstance(item, Worker):
            self.items.append(item)

    def remove_item(self):
        if not self.is_empty():
            return self.items.pop(0)
        return None


class Storage(Place):
    def __init__(self, name):
        super().__init__(name)

    def add_item(self, item):
        if isinstance(item, Product):
            self.items.append(item)

    def remove_item(self):
        if not self.is_empty():
            return self.items.pop()
        return None


class FoodStorage(Place):
    def __init__(self, name):
        super().__init__(name)

    def add_item(self, item):
        if isinstance(item, Food):
            self.items.append(item)

    def remove_item(self):
        if not self.is_empty():
            return self.items.pop()
        return None


class Transition:
    def __init__(self, name):
        self.name = name

    def can_fire(self):
        return False

    def fire(self):
        return False

    def __repr__(self):
        return f"Transition: {self.name}"


class Factory(Transition):
    def __init__(
        self, name, worker_input, worker_output, product_output, danger_level=15
    ):
        super().__init__(name)
        self.worker_input = worker_input
        self.worker_output = worker_output
        self.product_output = product_output
        self.danger_level = danger_level

    def can_fire(self):
        return not self.worker_input.is_empty()

    def fire(self):
        if not self.can_fire():
            return False

        worker = self.worker_input.remove_item()


        damage = random.randint(0, self.danger_level)
        worker.adjust_life_force(-damage)

        if worker.is_alive:
            product = Product(get_next_product_id())
            self.product_output.add_item(product)
            print(f"  {worker} produced {product} at {self.name}")
            self.worker_output.add_item(worker)
            return True
        else:
            print(f"  {worker} died at {self.name}")
            return False


class Cafeteria(Transition):
    def __init__(self, name, worker_input, food_input, worker_output):
        super().__init__(name)
        self.worker_input = worker_input
        self.food_input = food_input
        self.worker_output = worker_output

    def can_fire(self):
        return not self.worker_input.is_empty() and not self.food_input.is_empty()

    def fire(self):
        if not self.can_fire():
            return False

        worker = self.worker_input.remove_item()
        food = self.food_input.remove_item()

        worker.adjust_life_force(food.nutrition_value)
        print(f"  {worker} ate {food} at {self.name}")

        self.worker_output.add_item(worker)
        return True


class Field(Transition):
    def __init__(self, name, worker_input, worker_output, food_output, danger_level=10):
        super().__init__(name)
        self.worker_input = worker_input
        self.worker_output = worker_output
        self.food_output = food_output
        self.danger_level = danger_level

    def can_fire(self):
        return not self.worker_input.is_empty()

    def fire(self):
        if not self.can_fire():
            return False

        worker = self.worker_input.remove_item()

        damage = random.randint(0, self.danger_level)
        worker.adjust_life_force(-damage)

        if worker.is_alive:
            food = Food(get_next_food_id(), nutrition_value=30)
            self.food_output.add_item(food)
            print(f"  {worker} produced {food} at {self.name}")
            self.worker_output.add_item(worker)
            return True
        else:
            print(f"  {worker} died at {self.name}")
            return False


class Home(Transition):
    def __init__(self, name, worker_input, product_input, worker_output):
        super().__init__(name)
        self.worker_input = worker_input
        self.worker_output = worker_output
        self.product_input = product_input

    def can_fire(self):
        return not self.worker_input.is_empty()

    def fire(self):
        if not self.can_fire():
            return False

        worker1 = self.worker_input.remove_item()
        self.product_input.remove_item()

        if not self.worker_input.is_empty():
            worker2 = self.worker_input.remove_item()

            baby = Worker(get_next_worker_id(), life_force=50)
            print(f"  {worker1} and {worker2} created {baby} at {self.name}")

            self.worker_output.add_item(worker1)
            self.worker_output.add_item(worker2)
            self.worker_output.add_item(baby)
            return True
        else:
            worker1.adjust_life_force(20)
            print(f"  {worker1} rested at {self.name}")
            self.worker_output.add_item(worker1)
            return True


class SimpleTransition(Transition):
    def __init__(self, name, input_place, output_place, danger_level=5):
        super().__init__(name)
        self.input_place = input_place
        self.output_place = output_place
        self.danger_level = danger_level

    def can_fire(self):
        return not self.input_place.is_empty()

    def fire(self):
        if not self.can_fire():
            return False

        item = self.input_place.remove_item()

        if isinstance(item, Worker):
            damage = random.randint(0, self.danger_level)
            item.adjust_life_force(-damage)

            if item.is_alive:
                self.output_place.add_item(item)
                return True
            else:
                print(f"  {item} died during {self.name}")
                return False
        else:
            self.output_place.add_item(item)
            return True


class Network:
    def __init__(self):
        self.places = {}
        self.transitions = []
        self.worker_history = []
        self.product_history = []
        self.food_history = []

    def plot_results(self):
        plt.figure()
        plt.plot(self.worker_history, label="Workers")
        plt.plot(self.product_history, label="Products")
        plt.plot(self.food_history, label="Foods")

        plt.xlabel('Iteration')
        plt.ylabel('Amount')
        plt.legend()

        save_path = 'sim-sims.png'
        plt.savefig(save_path)
        plt.close()

        if platform.system() == 'Windows':
            os.startfile(save_path)


    def save_to_excel(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Simulation Data'

        headers = ['Iteration', 'Workers', 'Products', 'Food']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, name='Times New Roman')
            cell.fill = PatternFill("solid", start_color="4F81BD")
            cell.alignment = Alignment(horizontal="center")

        for i, (w, p, f) in enumerate(
            zip(self.worker_history, self.product_history, self.food_history)
        ):
            sheet.append([i, w, p,f])

        for col in ["A", "B", "C", "D"]:
            sheet.column_dimensions[col].width = 15

        wb.save("simulation_results.xlsx")
        print("Data saved to simulation_results.xlsx")

        if platform.system() == 'Windows':
            os.startfile('simulation_results.xlsx')


    def collect_statistics(self):
        total_workers = 0
        total_products = 0
        total_foods = 0

        for place in self.places.values():
            for item in place.items:
                if isinstance(item, Worker):
                    total_workers += 1
                elif isinstance(item, Product):
                    total_products += 1
                elif isinstance(item, Food):
                    total_foods += 1
        self.worker_history.append(total_workers)
        self.product_history.append(total_products)
        self.food_history.append(total_foods)



    def add_place(self, place):
        self.places[place.name] = place

    def add_transition(self, transition):
        self.transitions.append(transition)

    def get_place(self, name):
        return self.places.get(name)

    def has_workers_in_system(self):
        for place in self.places.values():
            if isinstance(place, Barracks) and not place.is_empty():
                return True
        return False

    def run(self, max_iterations=1000):
        iteration = 0

        print("=== Initial State ===")
        self.print_state()
        self.collect_statistics()
        print()

        while self.has_workers_in_system() and iteration < max_iterations:
            fired_any = False

            for transition in self.transitions:
                if transition.fire():
                    fired_any = True
                    print(f"Transition '{transition.name}' fired\n")
                    self.print_state()
                    time.sleep(0.1)

            self.collect_statistics()

            if not fired_any:
                print("\nNo more transitions can be activated")
                break

            iteration += 1

        if iteration >= max_iterations:
            print(f"\nReached max iterations ({max_iterations})")

        if not self.has_workers_in_system():
            print("\nNo workers left in barracks")

        print("\n=== Final State ===")
        self.print_state()
        print(f"\nTotal iterations: {iteration}")

    def print_state(self):
        print("\nBarracks:")
        for name, place in self.places.items():
            if isinstance(place, Barracks):
                print(f"  {place}")

        print("\nStorage:")
        for name, place in self.places.items():
            if isinstance(place, Storage):
                print(f"  {place}")

        print("\nFood Storage:")
        for name, place in self.places.items():
            if isinstance(place, FoodStorage):
                print(f"  {place}")


def main():
    global next_worker_id, next_food_id, next_product_id
    next_worker_id = 1
    next_food_id = 1
    next_product_id = 1

    network = Network()

    barracks1 = Barracks("Barracks 1")
    barracks2 = Barracks("Barracks 2")
    barracks3 = Barracks("Barracks 3")
    barracks4 = Barracks("Barracks 4")
    barracks5 = Barracks("Barracks 5")

    storage1 = Storage("Storage 1")
    storage2 = Storage("Storage 2")

    food_storage1 = FoodStorage("Food Storage 1")
    food_storage2 = FoodStorage("Food Storage 2")

    network.add_place(barracks1)
    network.add_place(barracks2)
    network.add_place(barracks3)
    network.add_place(barracks4)
    network.add_place(barracks5)
    network.add_place(storage1)
    network.add_place(storage2)
    network.add_place(food_storage1)
    network.add_place(food_storage2)

    t1 = SimpleTransition("Move B1 to B2", barracks1, barracks2, danger_level=90)
    field1 = Field("Field Work", barracks2, barracks3, food_storage1, danger_level=80)
    cafeteria1 = Cafeteria("Lunch", barracks3, food_storage1, barracks4)
    factory1 = Factory("Production", barracks4, barracks5, storage1, danger_level=80)
    home1 = Home('Rest or Birth', barracks5, storage1, barracks1)

    network.add_transition(t1)
    network.add_transition(field1)
    network.add_transition(cafeteria1)
    network.add_transition(factory1)
    network.add_transition(home1)

    print("\nAdding initial workers to Barracks 1...")
    for i in range(1, 4):
        worker = Worker(get_next_worker_id())
        barracks1.add_item(worker)

    print("Adding initial workers to Barracks 3...")
    for i in range(1, 3):
        worker = Worker(get_next_worker_id())
        barracks3.add_item(worker)

    print("Adding initial food to Food Storage 1...")
    for i in range(1, 10):
        food = Food(get_next_food_id())
        food_storage1.add_item(food)

    total_count = len(network.places) + len(network.transitions)
    print(f"\nTotal places and transitions: {total_count}")
    if total_count >= 11:
        print("Requirement met (at least 11)")
    else:
        print(f"WARNING: Need at least 11, have {total_count}")

    print("\n" + "=" * 70)
    print("STARTING SIMULATION")
    print("=" * 70)

    network.run()
    network.plot_results()
    network.save_to_excel()



if __name__ == "__main__":
    main()
