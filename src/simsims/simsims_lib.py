import time
import os
import random
import matplotlib.pyplot as plt
import platform
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from simsims import id_factory


class Worker:
    """Represents a worker in the simulation.

    :param worker_id: Unique identifier for the worker.
    :type worker_id: int
    :param life_force: Starting health points of the worker, defaults to 100.
    :type life_force: int
    """

    def __init__(self, worker_id, life_force=100):
        """Initialize a Worker with an ID and life force."""
        self.worker_id = worker_id
        self.life_force = life_force
        self.is_alive = True

    def adjust_life_force(self, amount):
        """Adjust the worker's life force by a given amount.

        Caps life force at 100. If life force drops to 0 or below,
        the worker is marked as dead.

        :param amount: Amount to add (positive) or subtract (negative).
        :type amount: int
        """
        self.life_force += amount
        if self.life_force > 100:
            self.life_force = 100
        if self.life_force <= 0:
            self.is_alive = False

    def __repr__(self):
        if self.is_alive:
            return f"Worker-: {self.worker_id}(HP:{self.life_force})"
        else:
            return f"Worker-: {self.worker_id}(DEAD)"


class Food:
    """Represents a food item in the simulation.

    :param food_id: Unique identifier for the food.
    :type food_id: int
    :param nutrition_value: HP restored when consumed, defaults to 30.
    :type nutrition_value: int
    """

    def __init__(self, food_id, nutrition_value=30):
        """Initialize a Food item with an ID and nutrition value."""
        self.food_id = food_id
        self.nutrition_value = nutrition_value

    def __repr__(self):
        return f"Food-{self.food_id}(+{self.nutrition_value}HP)"


class Product:
    """Represents a product created by workers in the simulation.

    :param product_id: Unique identifier for the product.
    :type product_id: int
    """

    def __init__(self, product_id):
        """Initialize a Product with an ID."""
        self.product_id = product_id

    def __repr__(self):
        return f"Product-{self.product_id}"


class Place:
    """Base class representing a place (node) in the Petri net.

    :param name: Name of the place.
    :type name: str
    """

    def __init__(self, name):
        """Initialize a Place with a name and an empty items list."""
        self.name = name
        self.items = []

    def is_empty(self):
        """Check whether the place contains no items.

        :return: True if empty, False otherwise.
        :rtype: bool
        """
        return len(self.items) == 0

    def add_item(self, item):
        """Add an item to the place.

        :param item: The item to add.
        """
        self.items.append(item)

    def remove_item(self):
        """Remove and return the first item from the place.

        :return: The removed item, or None if empty.
        """
        if not self.is_empty():
            return self.items.pop(0)
        return None

    def __repr__(self):
        return f"{self.name}: {self.items}"


class Barracks(Place):
    """A place that holds Worker objects.

    :param name: Name of the barracks.
    :type name: str
    """

    def __init__(self, name):
        """Initialize Barracks by calling the parent Place constructor."""
        super().__init__(name)

    def add_item(self, item):
        """Add a Worker to the barracks. Non-Worker items are ignored.

        :param item: The item to add (must be a Worker).
        """
        if isinstance(item, Worker):
            self.items.append(item)

    def remove_item(self):
        """Remove and return the first Worker from the barracks.

        :return: The removed Worker, or None if empty.
        :rtype: Worker or None
        """
        if not self.is_empty():
            return self.items.pop(0)
        return None


class Storage(Place):
    """A place that holds Product objects.

    :param name: Name of the storage.
    :type name: str
    """

    def __init__(self, name):
        """Initialize Storage by calling the parent Place constructor."""
        super().__init__(name)

    def add_item(self, item):
        """Add a Product to storage. Non-Product items are ignored.

        :param item: The item to add (must be a Product).
        """
        if isinstance(item, Product):
            self.items.append(item)

    def remove_item(self):
        """Remove and return the last Product from storage.

        :return: The removed Product, or None if empty.
        :rtype: Product or None
        """
        if not self.is_empty():
            return self.items.pop()
        return None


class FoodStorage(Place):
    """A place that holds Food objects.

    :param name: Name of the food storage.
    :type name: str
    """

    def __init__(self, name):
        """Initialize FoodStorage by calling the parent Place constructor."""
        super().__init__(name)

    def add_item(self, item):
        """Add a Food item to storage. Non-Food items are ignored.

        :param item: The item to add (must be a Food).
        """
        if isinstance(item, Food):
            self.items.append(item)

    def remove_item(self):
        """Remove and return the last Food item from storage.

        :return: The removed Food, or None if empty.
        :rtype: Food or None
        """
        if not self.is_empty():
            return self.items.pop()
        return None


class Transition:
    """Base class representing a transition in the Petri net.

    :param name: Name of the transition.
    :type name: str
    """

    def __init__(self, name):
        """Initialize a Transition with a name."""
        self.name = name

    def can_fire(self):
        """Check whether this transition can fire.

        :return: Always False in the base class.
        :rtype: bool
        """
        return False

    def fire(self):
        """Attempt to fire this transition.

        :return: Always False in the base class.
        :rtype: bool
        """
        return False

    def __repr__(self):
        return f"Transition: {self.name}"


class Factory(Transition):
    """A transition where workers produce products, with a risk of injury or death.

    :param name: Name of the factory transition.
    :type name: str
    :param worker_input: Barracks providing workers.
    :type worker_input: Barracks
    :param worker_output: Barracks receiving workers after production.
    :type worker_output: Barracks
    :param product_output: Storage receiving produced products.
    :type product_output: Storage
    :param danger_level: Maximum possible damage per firing, defaults to 15.
    :type danger_level: int
    """

    def __init__(self, name, worker_input, worker_output, product_output, danger_level=15):
        """Initialize the Factory transition."""
        super().__init__(name)
        self.worker_input = worker_input
        self.worker_output = worker_output
        self.product_output = product_output
        self.danger_level = danger_level

    def can_fire(self):
        """Check if there is a worker available to work.

        :return: True if worker_input is not empty.
        :rtype: bool
        """
        return not self.worker_input.is_empty()

    def fire(self):
        """Fire the transition: a worker produces a product and may take damage.

        :return: True if the worker survived, False otherwise.
        :rtype: bool
        """
        if not self.can_fire():
            return False

        worker = self.worker_input.remove_item()
        damage = random.randint(0, self.danger_level)
        worker.adjust_life_force(-damage)

        if worker.is_alive:
            id = id_factory.get_next_product_id()
            product = Product(id)
            self.product_output.add_item(product)
            print(f"  {worker} produced {product} at {self.name}")
            self.worker_output.add_item(worker)
            return True
        else:
            print(f"  {worker} died at {self.name}")
            return False


class Cafeteria(Transition):
    """A transition where workers eat food to restore life force.

    :param name: Name of the cafeteria transition.
    :type name: str
    :param worker_input: Barracks providing hungry workers.
    :type worker_input: Barracks
    :param food_input: FoodStorage providing food.
    :type food_input: FoodStorage
    :param worker_output: Barracks receiving fed workers.
    :type worker_output: Barracks
    """

    def __init__(self, name, worker_input, food_input, worker_output):
        """Initialize the Cafeteria transition."""
        super().__init__(name)
        self.worker_input = worker_input
        self.food_input = food_input
        self.worker_output = worker_output

    def can_fire(self):
        """Check if both a worker and food are available.

        :return: True if both worker_input and food_input are non-empty.
        :rtype: bool
        """
        return not self.worker_input.is_empty() and not self.food_input.is_empty()

    def fire(self):
        """Fire the transition: a worker eats food and gains life force.

        :return: True if the transition fired successfully.
        :rtype: bool
        """
        if not self.can_fire():
            return False

        worker = self.worker_input.remove_item()
        food = self.food_input.remove_item()

        worker.adjust_life_force(food.nutrition_value)
        print(f"  {worker} ate {food} at {self.name}")

        self.worker_output.add_item(worker)
        return True


class Field(Transition):
    """A transition where workers harvest food, with a risk of injury or death.

    :param name: Name of the field transition.
    :type name: str
    :param worker_input: Barracks providing workers.
    :type worker_input: Barracks
    :param worker_output: Barracks receiving workers after harvesting.
    :type worker_output: Barracks
    :param food_output: FoodStorage receiving harvested food.
    :type food_output: FoodStorage
    :param danger_level: Maximum possible damage per firing, defaults to 10.
    :type danger_level: int
    """

    def __init__(self, name, worker_input, worker_output, food_output, danger_level=10):
        """Initialize the Field transition."""
        super().__init__(name)
        self.worker_input = worker_input
        self.worker_output = worker_output
        self.food_output = food_output
        self.danger_level = danger_level

    def can_fire(self):
        """Check if there is a worker available to harvest.

        :return: True if worker_input is not empty.
        :rtype: bool
        """
        return not self.worker_input.is_empty()

    def fire(self):
        """Fire the transition: a worker harvests food and may take damage.

        :return: True if the worker survived, False otherwise.
        :rtype: bool
        """
        if not self.can_fire():
            return False

        worker = self.worker_input.remove_item()
        damage = random.randint(0, self.danger_level)
        worker.adjust_life_force(-damage)

        if worker.is_alive:
            food = Food(id_factory.get_next_food_id(), nutrition_value=30)
            self.food_output.add_item(food)
            print(f"  {worker} produced {food} at {self.name}")
            self.worker_output.add_item(worker)
            return True
        else:
            print(f"  {worker} died at {self.name}")
            return False


class Home(Transition):
    """A transition where workers rest or reproduce to create new workers.

    If two workers are available, they create a baby worker.
    If only one worker is present, they rest and recover life force.

    :param name: Name of the home transition.
    :type name: str
    :param worker_input: Barracks providing workers.
    :type worker_input: Barracks
    :param product_input: Storage providing products (consumed during reproduction).
    :type product_input: Storage
    :param worker_output: Barracks receiving workers (and new baby) after the transition.
    :type worker_output: Barracks
    """

    def __init__(self, name, worker_input, product_input, worker_output):
        """Initialize the Home transition."""
        super().__init__(name)
        self.worker_input = worker_input
        self.worker_output = worker_output
        self.product_input = product_input

    def can_fire(self):
        """Check if at least one worker is available.

        :return: True if worker_input is not empty.
        :rtype: bool
        """
        return not self.worker_input.is_empty()

    def fire(self):
        """Fire the transition: workers rest or reproduce.

        :return: True if the transition fired successfully.
        :rtype: bool
        """
        if not self.can_fire():
            return False

        worker1 = self.worker_input.remove_item()
        self.product_input.remove_item()

        if not self.worker_input.is_empty():
            worker2 = self.worker_input.remove_item()
            baby = Worker(id_factory.get_next_worker_id(), life_force=50)
            print(f"  {worker1} and {worker2} created {baby} at {self.name}")
            self.worker_output.add_item(worker1)
            self.worker_output.add_item(worker2)
            self.worker_output.add_item(baby)
            return True
        else:
            worker1.adjust_life_force(20)
            print(f"  {worker1} rested at {self.name}")
            self.worker_output.add_item(worker1)
            return True


class SimpleTransition(Transition):
    """A basic transition that moves an item from one place to another.

    Workers may take random damage during the move.

    :param name: Name of the transition.
    :type name: str
    :param input_place: Place to take the item from.
    :type input_place: Place
    :param output_place: Place to send the item to.
    :type output_place: Place
    :param danger_level: Maximum possible damage if item is a Worker, defaults to 5.
    :type danger_level: int
    """

    def __init__(self, name, input_place, output_place, danger_level=5):
        """Initialize the SimpleTransition."""
        super().__init__(name)
        self.input_place = input_place
        self.output_place = output_place
        self.danger_level = danger_level

    def can_fire(self):
        """Check if the input place has an item.

        :return: True if input_place is not empty.
        :rtype: bool
        """
        return not self.input_place.is_empty()

    def fire(self):
        """Fire the transition: move an item, applying damage if it is a Worker.

        :return: True if item was moved successfully, False if Worker died.
        :rtype: bool
        """
        if not self.can_fire():
            return False

        item = self.input_place.remove_item()

        if isinstance(item, Worker):
            damage = random.randint(0, self.danger_level)
            item.adjust_life_force(-damage)

            if item.is_alive:
                self.output_place.add_item(item)
                return True
            else:
                print(f"  {item} died during {self.name}")
                return False
        else:
            self.output_place.add_item(item)
            return True


class Network:
    """Represents the full Petri net simulation network.

    Manages all places, transitions, and simulation history.
    """

    def __init__(self):
        """Initialize an empty Network with no places, transitions, or history."""
        self.places = {}
        self.transitions = []
        self.worker_history = []
        self.product_history = []
        self.food_history = []

    def plot_results(self):
        """Plot the simulation history of workers, products, and food over iterations.

        Saves the plot as 'sim-sims.png' and opens it on Windows.
        """
        plt.figure()
        plt.plot(self.worker_history, label="Workers")
        plt.plot(self.product_history, label="Products")
        plt.plot(self.food_history, label="Foods")

        plt.xlabel('Iteration')
        plt.ylabel('Amount')
        plt.legend()

        save_path = 'sim-sims.png'
        plt.savefig(save_path)
        plt.close()

        if platform.system() == 'Windows':
            os.startfile(save_path)

    def save_to_excel(self):
        """Save the simulation history to an Excel file named 'simulation_results.xlsx'.

        Includes columns for iteration, worker count, product count, and food count.
        Opens the file automatically on Windows.
        """
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Simulation Data'

        headers = ['Iteration', 'Workers', 'Products', 'Food']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, name='Times New Roman')
            cell.fill = PatternFill("solid", start_color="4F81BD")
            cell.alignment = Alignment(horizontal="center")

        for i, (w, p, f) in enumerate(
            zip(self.worker_history, self.product_history, self.food_history)
        ):
            sheet.append([i, w, p, f])

        for col in ["A", "B", "C", "D"]:
            sheet.column_dimensions[col].width = 15

        wb.save("simulation_results.xlsx")
        print("Data saved to simulation_results.xlsx")

        if platform.system() == 'Windows':
            os.startfile('simulation_results.xlsx')

    def collect_statistics(self):
        """Count and record the current number of workers, products, and food items.

        Appends counts to the respective history lists.
        """
        total_workers = 0
        total_products = 0
        total_foods = 0

        for place in self.places.values():
            for item in place.items:
                if isinstance(item, Worker):
                    total_workers += 1
                elif isinstance(item, Product):
                    total_products += 1
                elif isinstance(item, Food):
                    total_foods += 1
        self.worker_history.append(total_workers)
        self.product_history.append(total_products)
        self.food_history.append(total_foods)

    def add_place(self, place):
        """Add a place to the network.

        :param place: The place to add.
        :type place: Place
        """
        self.places[place.name] = place

    def add_transition(self, transition):
        """Add a transition to the network.

        :param transition: The transition to add.
        :type transition: Transition
        """
        self.transitions.append(transition)

    def get_place(self, name):
        """Retrieve a place by name.

        :param name: The name of the place.
        :type name: str
        :return: The matching Place, or None if not found.
        :rtype: Place or None
        """
        return self.places.get(name)

    def has_workers_in_system(self):
        """Check whether any Barracks in the network still contains workers.

        :return: True if at least one Barracks is non-empty.
        :rtype: bool
        """
        for place in self.places.values():
            if isinstance(place, Barracks) and not place.is_empty():
                return True
        return False

    def run(self, max_iterations=1000):
        """Run the simulation until no workers remain or max iterations is reached.

        :param max_iterations: Maximum number of iterations before stopping, defaults to 1000.
        :type max_iterations: int
        """
        iteration = 0

        print("=== Initial State ===")
        self.print_state()
        self.collect_statistics()
        print()

        while self.has_workers_in_system() and iteration < max_iterations:
            fired_any = False

            for transition in self.transitions:
                if transition.fire():
                    fired_any = True
                    print(f"Transition '{transition.name}' fired\n")
                    self.print_state()
                    time.sleep(0.1)

            self.collect_statistics()

            if not fired_any:
                print("\nNo more transitions can be activated")
                break

            iteration += 1

        if iteration >= max_iterations:
            print(f"\nReached max iterations ({max_iterations})")

        if not self.has_workers_in_system():
            print("\nNo workers left in barracks")

        print("\n=== Final State ===")
        self.print_state()
        print(f"\nTotal iterations: {iteration}")

    def print_state(self):
        """Print the current state of all Barracks, Storage, and FoodStorage places."""
        print("\nBarracks:")
        for name, place in self.places.items():
            if isinstance(place, Barracks):
                print(f"  {place}")

        print("\nStorage:")
        for name, place in self.places.items():
            if isinstance(place, Storage):
                print(f"  {place}")

        print("\nFood Storage:")
        for name, place in self.places.items():
            if isinstance(place, FoodStorage):
                print(f"  {place}")
